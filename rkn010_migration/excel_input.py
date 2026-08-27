from __future__ import annotations

import re
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from .models import SourceRow, ValidationIssue, WorkbookData


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "row_uuid": ("row_uuid",),
    "source_record_id": ("идентификатор записи",),
    "regno": ("регистрационный номер",),
    "ogrn": ("огрн",),
    "org_name": ("полное наименование",),
    "short_org_name": ("сокращенное наименование", "сокращенное"),
    "location": ("место нахождения",),
    "licence_number_1": ("n1", "licencenumber1", "licence number 1"),
    "licence_number_2": ("n2", "licencenumber2", "licence number 2"),
    "licence_number_3": ("n3", "licencenumber3", "licence number 3"),
    "geo_zone": ("зона нумерации", "зона", "geo_zone"),
    "include_order_number": ("номер приказа о включении", "№ приказа о включении"),
    "include_order_date": ("дата приказа о включении",),
    "exclude_order_number": ("номер приказа об исключении", "№ приказа об исключении"),
    "exclude_order_date": ("дата приказа об исключении",),
    "expert_opinion_date": ("дата экспертного заключения",),
}

HEADER_PATTERNS: dict[str, str] = {
    "regno": r"^регистрацион.*номер",
    "org_name": r"^полное наименование оператора",
    "short_org_name": r"^сокращенное наименование оператора",
    "location": r"^место нахождения",
    "licence_number_1": r"^номера лицензий n1",
    "licence_number_2": r"^номера лицензий n2",
    "licence_number_3": r"^номера лицензий n3",
    "geo_zone": r"^географически определенная зона нумерации",
    "include_order_number": r"^№ приказа о включении в реестр",
    "include_order_date": r"^дата приказа о включении в реестр",
    "exclude_order_number": r"^№ приказа об исключении из реестра",
    "exclude_order_date": r"^дата приказа об исключении из реестра",
}

REQUIRED_FIELDS = {
    "regno",
    "ogrn",
    "org_name",
    "location",
    "geo_zone",
    "include_order_number",
    "include_order_date",
}


class WorkbookValidationError(ValueError):
    def __init__(self, issues: list[ValidationIssue]):
        self.issues = issues
        text = "\n".join(
            f"row {issue.excel_row or '-'} [{issue.code}] {issue.message}"
            for issue in issues
            if issue.severity == "error"
        )
        super().__init__(text or "Workbook validation failed")


def _normalise_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"[\s._-]+", " ", text)
    return text.strip()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date(value: Any, *, epoch: datetime) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        result = value
    elif isinstance(value, date):
        result = datetime.combine(value, time.min)
    elif isinstance(value, (int, float)):
        result = from_excel(value, epoch=epoch)
        if isinstance(result, time):
            raise ValueError("time without date")
        if isinstance(result, date) and not isinstance(result, datetime):
            result = datetime.combine(result, time.min)
    else:
        text = _text(value)
        result = None
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
            try:
                result = datetime.strptime(text[:10], fmt)
                break
            except ValueError:
                continue
        if result is None:
            try:
                result = datetime.fromisoformat(text.replace("Z", "+00:00"))
            except ValueError as exc:
                raise ValueError(f"unsupported date {text!r}") from exc
    if result.tzinfo is not None:
        result = result.astimezone(timezone.utc).replace(tzinfo=None)
    return result.replace(hour=0, minute=0, second=0, microsecond=0)


def _find_header(ws) -> tuple[int, dict[str, int]]:
    aliases = {
        field: {_normalise_header(alias) for alias in values}
        for field, values in HEADER_ALIASES.items()
    }
    best: tuple[int, dict[str, int]] | None = None
    for row_number in range(1, min(ws.max_row, 20) + 1):
        found: dict[str, int] = {}
        for column in range(1, ws.max_column + 1):
            header = _normalise_header(ws.cell(row_number, column).value)
            for field, choices in aliases.items():
                pattern = HEADER_PATTERNS.get(field)
                if header in choices or (pattern and re.search(pattern, header)):
                    found[field] = column
        if best is None or len(found) > len(best[1]):
            best = (row_number, found)
    if best is None:
        raise WorkbookValidationError(
            [ValidationIssue("header_not_found", "Не удалось найти строку заголовков")]
        )
    missing = sorted(set(HEADER_ALIASES) - set(best[1]))
    if missing:
        raise WorkbookValidationError(
            [ValidationIssue("missing_columns", f"Отсутствуют столбцы: {', '.join(missing)}")]
        )
    return best


def read_workbook(path: str | Path, sheet_name: str | None = None) -> WorkbookData:
    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    # Some source exports do not persist worksheet dimensions. In read-only
    # mode openpyxl then exposes max_row/max_column as None, so use normal mode.
    # The RKN010 workbook is small enough that this has negligible memory cost.
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb["РОСЗСП"] if "РОСЗСП" in wb.sheetnames else wb.active
    header_row, columns = _find_header(ws)
    rows: list[SourceRow] = []
    issues: list[ValidationIssue] = []

    for row_number in range(header_row + 1, ws.max_row + 1):
        raw = {field: ws.cell(row_number, col).value for field, col in columns.items()}
        if not any(value not in (None, "") for value in raw.values()):
            continue
        values = {field: _text(value) for field, value in raw.items()}
        for field in REQUIRED_FIELDS:
            if not values[field]:
                issues.append(ValidationIssue("required", f"Пустое обязательное поле {field}", row_number))
        if values["ogrn"] and not re.fullmatch(r"\d{13}|\d{15}", values["ogrn"]):
            issues.append(ValidationIssue("ogrn_format", "ОГРН должен содержать 13 или 15 цифр", row_number))
        if values["geo_zone"] and not re.fullmatch(r"\d+", values["geo_zone"]):
            issues.append(ValidationIssue("geo_zone_format", "В строке должна быть ровно одна цифровая зона", row_number))
        if not any(values[name] for name in ("licence_number_1", "licence_number_2", "licence_number_3")):
            issues.append(ValidationIssue("licence_numbers_empty", "Не заполнен ни один номер N1/N2/N3", row_number))
        try:
            include_date = _date(raw["include_order_date"], epoch=wb.epoch)
        except ValueError as exc:
            issues.append(ValidationIssue("include_date_format", str(exc), row_number))
            include_date = None
        try:
            exclude_date = _date(raw["exclude_order_date"], epoch=wb.epoch)
        except ValueError as exc:
            issues.append(ValidationIssue("exclude_date_format", str(exc), row_number))
            exclude_date = None
        try:
            opinion_date = _date(raw["expert_opinion_date"], epoch=wb.epoch)
        except ValueError as exc:
            issues.append(ValidationIssue("opinion_date_format", str(exc), row_number))
            opinion_date = None
        if bool(values["exclude_order_number"]) != bool(exclude_date):
            issues.append(ValidationIssue("cancellation_pair", "Номер и дата приказа об исключении должны быть заполнены вместе", row_number))
        rows.append(
            SourceRow(
                excel_row=row_number,
                row_uuid=values["row_uuid"],
                source_record_id=values["source_record_id"],
                regno=values["regno"],
                ogrn=values["ogrn"],
                org_name=values["org_name"],
                short_org_name=values["short_org_name"],
                location=values["location"],
                licence_number_1=values["licence_number_1"],
                licence_number_2=values["licence_number_2"],
                licence_number_3=values["licence_number_3"],
                geo_zone=values["geo_zone"],
                include_order_number=values["include_order_number"],
                include_order_date=include_date,
                exclude_order_number=values["exclude_order_number"],
                exclude_order_date=exclude_date,
                expert_opinion_date=opinion_date,
            )
        )

    if not rows:
        issues.append(ValidationIssue("no_rows", "В таблице нет строк данных"))
    actual_sheet_name = ws.title
    wb.close()
    return WorkbookData(str(workbook_path), actual_sheet_name, rows, issues)


def raise_for_errors(data: WorkbookData) -> None:
    if data.errors:
        raise WorkbookValidationError(data.errors)
